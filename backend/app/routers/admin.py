from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import Dict, Any, List, Optional
import io
import csv
import openpyxl

from app.dependencies import get_db, RoleRequired
from app.models.user import User
from app.models.student import Student, FacultyProfile, MentorAssignment
from app.models.score import StudentAnalytics
from app.models.profile import UserProfile, StudentAbout
from app.models.portfolio import PortfolioCustomization
from app.utils.security import get_password_hash
from app.services.analytics_service import recalculate_student_analytics
from app.schemas.score import UploadScoresResponse

router = APIRouter()

@router.get("/overview")
async def get_admin_overview():
    """Skeleton route for retrieving administrative system overview statistics."""
    return {
        "total_students": 0,
        "total_faculty": 0,
        "total_mentors": 0,
        "total_users": 0,
        "total_scores_uploaded": 0
    }

@router.get("/counts")
async def get_admin_counts(
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """Retrieves lightweight administrative counts for students, faculty, mentors, and scores."""
    from app.models.student import Student
    from app.models.score import AssessmentScore
    
    total_students = db.query(Student).count()
    total_faculty = db.query(User).filter(User.role == "faculty").count()
    total_mentors = db.query(User).filter(User.role == "mentor").count()
    total_scores = db.query(AssessmentScore).count()
    
    return {
        "students": total_students,
        "faculty": total_faculty,
        "mentors": total_mentors,
        "scores": total_scores
    }

# Students CRUD
@router.get("/students")
async def admin_get_students(
    page: Optional[int] = None,
    limit: Optional[int] = None,
    search: Optional[str] = None,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    from sqlalchemy.orm import joinedload
    query = db.query(Student).options(joinedload(Student.user))
    
    if search:
        query = query.filter(
            (Student.register_no.ilike(f"%{search}%")) |
            (Student.name.ilike(f"%{search}%")) |
            (Student.email.ilike(f"%{search}%"))
        )
    
    total = query.count()
    
    if page is not None and limit is not None:
        offset = (page - 1) * limit
        students = query.offset(offset).limit(limit).all()
        result = []
        for s in students:
            s_data = {
                "id": s.id,
                "register_no": s.register_no,
                "name": s.name,
                "email": s.email,
                "phone": s.phone,
                "department": s.department,
                "year": s.year,
                "section": s.section,
                "batch": s.batch,
                "user_id": s.user_id,
            }
            if s.user:
                s_data["user_email"] = s.user.email
                s_data["role"] = s.user.role
                s_data["is_active"] = s.user.is_active
                s_data["status"] = "Active" if s.user.is_active else "Inactive"
            else:
                s_data["user_email"] = None
                s_data["role"] = None
                s_data["is_active"] = None
                s_data["status"] = "Inactive"
            result.append(s_data)
        return {
            "items": result,
            "total": total,
            "page": page,
            "limit": limit
        }
    else:
        students = query.all()
        result = []
        for s in students:
            s_data = {
                "id": s.id,
                "register_no": s.register_no,
                "name": s.name,
                "email": s.email,
                "phone": s.phone,
                "department": s.department,
                "year": s.year,
                "section": s.section,
                "batch": s.batch,
                "user_id": s.user_id,
            }
            if s.user:
                s_data["user_email"] = s.user.email
                s_data["role"] = s.user.role
                s_data["is_active"] = s.user.is_active
                s_data["status"] = "Active" if s.user.is_active else "Inactive"
            else:
                s_data["user_email"] = None
                s_data["role"] = None
                s_data["is_active"] = None
                s_data["status"] = "Inactive"
            result.append(s_data)
        return result

# Faculty CRUD
@router.get("/faculty")
async def admin_get_faculty(
    page: Optional[int] = None,
    limit: Optional[int] = None,
    search: Optional[str] = None,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    from sqlalchemy.orm import joinedload
    query = db.query(User).filter(User.role == "faculty")
    
    if search:
        query = query.outerjoin(FacultyProfile).outerjoin(UserProfile).filter(
            (User.email.ilike(f"%{search}%")) |
            (User.username.ilike(f"%{search}%")) |
            (FacultyProfile.name.ilike(f"%{search}%")) |
            (UserProfile.full_name.ilike(f"%{search}%"))
        )
        
    total = query.count()
    query = query.options(joinedload(User.faculty_profile), joinedload(User.user_profile))
    
    if page is not None and limit is not None:
        offset = (page - 1) * limit
        users = query.offset(offset).limit(limit).all()
        result = []
        for u in users:
            fp = u.faculty_profile
            result.append({
                "id": u.id,
                "name": fp.name if fp else u.user_profile.full_name if u.user_profile else u.username,
                "email": u.email,
                "department": fp.department if fp else u.user_profile.department if u.user_profile else "CSE",
                "role": u.role,
                "phone": fp.phone if fp else u.user_profile.phone if u.user_profile else "",
                "status": "Active" if u.is_active else "Inactive",
                "designation": fp.designation if fp else "Faculty"
            })
        return {
            "items": result,
            "total": total,
            "page": page,
            "limit": limit
        }
    else:
        users = query.all()
        result = []
        for u in users:
            fp = u.faculty_profile
            result.append({
                "id": u.id,
                "name": fp.name if fp else u.user_profile.full_name if u.user_profile else u.username,
                "email": u.email,
                "department": fp.department if fp else u.user_profile.department if u.user_profile else "CSE",
                "role": u.role,
                "phone": fp.phone if fp else u.user_profile.phone if u.user_profile else "",
                "status": "Active" if u.is_active else "Inactive",
                "designation": fp.designation if fp else "Faculty"
            })
        return result

# Mentors CRUD
@router.get("/mentors")
async def admin_get_mentors(
    page: Optional[int] = None,
    limit: Optional[int] = None,
    search: Optional[str] = None,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    from sqlalchemy.orm import joinedload
    query = db.query(User).filter(User.role == "mentor")
    
    if search:
        query = query.outerjoin(FacultyProfile).outerjoin(UserProfile).filter(
            (User.email.ilike(f"%{search}%")) |
            (User.username.ilike(f"%{search}%")) |
            (FacultyProfile.name.ilike(f"%{search}%")) |
            (UserProfile.full_name.ilike(f"%{search}%"))
        )
        
    total = query.count()
    query = query.options(joinedload(User.faculty_profile), joinedload(User.user_profile))
    
    if page is not None and limit is not None:
        offset = (page - 1) * limit
        users = query.offset(offset).limit(limit).all()
    else:
        users = query.all()
        
    user_ids = [u.id for u in users]
    assignments = {}
    if user_ids:
        db_assignments = db.query(MentorAssignment).options(joinedload(MentorAssignment.student)).filter(
            MentorAssignment.mentor_id.in_(user_ids)
        ).all()
        for ass in db_assignments:
            if ass.mentor_id not in assignments:
                assignments[ass.mentor_id] = ass
                
    result = []
    for u in users:
        up = u.user_profile
        fp = u.faculty_profile
        assigned_class = "None"
        assignment = assignments.get(u.id)
        if assignment and assignment.student:
            s = assignment.student
            assigned_class = f"{s.year} {s.section} ({s.batch})"
            
        result.append({
            "id": u.id,
            "name": up.full_name if up else fp.name if fp else u.username,
            "email": u.email,
            "department": up.department if up else fp.department if fp else "IT",
            "mentorType": "Class Mentor",
            "assignedClass": assigned_class,
            "status": "Active" if u.is_active else "Inactive",
            "phone": up.phone if up else fp.phone if fp else ""
        })
        
    if page is not None and limit is not None:
        return {
            "items": result,
            "total": total,
            "page": page,
            "limit": limit
        }
    else:
        return result


@router.post("/mentors")
async def admin_create_mentor(payload: Dict[str, Any]):
    return {"success": True}

@router.put("/mentors/{id}")
async def admin_update_mentor(id: int, payload: Dict[str, Any]):
    return {"success": True}

@router.delete("/mentors/{id}")
async def admin_delete_mentor(id: int):
    return {"success": True}


# Users CRUD
@router.get("/users")
async def admin_get_users(
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    users = db.query(User).all()
    result = []
    for u in users:
        dept = None
        if u.student_profile and u.student_profile.department:
            dept = u.student_profile.department
        elif u.faculty_profile and u.faculty_profile.department:
            dept = u.faculty_profile.department
        elif u.user_profile and u.user_profile.department:
            dept = u.user_profile.department

        result.append({
            "id": u.id,
            "email": u.email,
            "username": u.username,
            "role": u.role,
            "is_active": u.is_active,
            "department": dept or "-",
            "name": u.name
        })
    return result

@router.post("/users")
async def admin_create_user(payload: Dict[str, Any]):
    return {"success": True}

@router.put("/users/{id}")
async def admin_update_user(id: int, payload: Dict[str, Any]):
    return {"success": True}

@router.delete("/users/{id}")
async def admin_delete_user(id: int):
    return {"success": True}


# Mentor Assignments CRUD
@router.get("/mentor-assignments")
async def admin_get_assignments():
    return []

@router.post("/mentor-assignments")
async def admin_create_assignment(payload: Dict[str, Any]):
    return {"success": True}

@router.delete("/mentor-assignments/{id}")
async def admin_delete_assignment(id: int):
    return {"success": True}


# Helper: Parse file to list of rows
def parse_upload_file(file: UploadFile) -> List[List[Any]]:
    contents = file.file.read()
    filename = file.filename.lower()
    
    if filename.endswith(".csv"):
        text_stream = io.StringIO(contents.decode("utf-8", errors="ignore"))
        reader = csv.reader(text_stream)
        return list(reader)
    else:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = workbook.active
            return list(sheet.iter_rows(values_only=True))
        except Exception as e:
            try:
                text_stream = io.StringIO(contents.decode("utf-8", errors="ignore"))
                reader = csv.reader(text_stream)
                return list(reader)
            except Exception:
                raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")


# Helper: Map headers
def map_headers(row_header: List[Any], expected_cols: Dict[str, List[str]]) -> Dict[str, int]:
    headers = [str(cell).strip().lower().replace("_", " ").replace("  ", " ") if cell is not None else "" for cell in row_header]
    col_mapping = {}
    for col_key, aliases in expected_cols.items():
        for alias in aliases:
            if alias in headers:
                col_mapping[col_key] = headers.index(alias)
                break
    return col_mapping


# Helper: Clean and generate unique email
def clean_and_generate_email(name: str, suffix_id: str, db: Session) -> str:
    cleaned_name = "".join(c.lower() for c in name if c.isalnum())
    email = f"{cleaned_name}@kce.ac.in"
    
    existing = db.query(User).filter(User.email == email).first()
    if existing:
        suffix = str(suffix_id).strip().lower()
        if suffix:
            email = f"{cleaned_name}{suffix}@kce.ac.in"
        else:
            count = 2
            while db.query(User).filter(User.email == f"{cleaned_name}{count}@kce.ac.in").first():
                count += 1
            email = f"{cleaned_name}{count}@kce.ac.in"
    return email


# Endpoint: Students Upload
@router.post("/upload/students")
async def upload_students(
    file: UploadFile = File(...),
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    rows = parse_upload_file(file)
    if not rows or len(rows) < 2:
        return {
            "success": False,
            "type": "students",
            "inserted": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [{"row": 1, "identifier": "File", "message": "Spreadsheet is empty or missing data rows"}]
        }
        
    expected_cols = {
        "register_no": ["register no", "register_no", "reg no", "reg_no", "register number", "reg number"],
        "name": ["name", "student name", "student_name", "full name", "full_name"],
        "department": ["department", "dept", "branch"],
        "year": ["year", "yr", "class year"],
        "section": ["section", "sec"],
        "email": ["email", "email address", "email_address"],
        "phone": ["phone", "phone no", "phone_no", "phone number", "mobile", "mobile no", "mobile_no"],
        "batch": ["batch", "year of passing", "grad year", "graduation year"],
        "mentor_email": ["mentor_email", "mentor email", "mentor_address", "mentor_email_address"],
        "date_of_birth": ["date_of_birth", "date of birth", "dob"],
        "gender": ["gender", "sex"],
        "address": ["address", "location", "residence"],
        "password": ["password", "default password", "default_password", "pass"]
    }
    
    col_mapping = map_headers(rows[0], expected_cols)
    for req in ["register_no", "name"]:
        if req not in col_mapping:
            raise HTTPException(status_code=400, detail=f"Missing required column: {req}")
            
    inserted = 0
    updated = 0
    skipped = 0
    errors_list = []
    
    hashed_default_password = get_password_hash("Password123!")
    
    for idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            skipped += 1
            continue
            
        try:
            reg_idx = col_mapping["register_no"]
            name_idx = col_mapping["name"]
            
            register_no = str(row[reg_idx]).strip() if (reg_idx < len(row) and row[reg_idx] is not None) else ""
            name = str(row[name_idx]).strip() if (name_idx < len(row) and row[name_idx] is not None) else ""
            
            if not register_no or register_no == "None":
                errors_list.append({"row": idx, "identifier": "Row " + str(idx), "message": "Register number is missing"})
                skipped += 1
                continue
            if not name or name == "None":
                errors_list.append({"row": idx, "identifier": register_no or "Row " + str(idx), "message": "Student name is missing"})
                skipped += 1
                continue
                
            dept_idx = col_mapping.get("department")
            department = str(row[dept_idx]).strip() if (dept_idx is not None and dept_idx < len(row) and row[dept_idx] is not None) else "AI & DS"
            
            year_idx = col_mapping.get("year")
            year = str(row[year_idx]).strip() if (year_idx is not None and year_idx < len(row) and row[year_idx] is not None) else "3"
            
            sec_idx = col_mapping.get("section")
            section = str(row[sec_idx]).strip() if (sec_idx is not None and sec_idx < len(row) and row[sec_idx] is not None) else "A"
            
            phone_idx = col_mapping.get("phone")
            phone = str(row[phone_idx]).strip() if (phone_idx is not None and phone_idx < len(row) and row[phone_idx] is not None) else ""
            
            batch_idx = col_mapping.get("batch")
            batch = str(row[batch_idx]).strip() if (batch_idx is not None and batch_idx < len(row) and row[batch_idx] is not None) else "2028"
            
            email_idx = col_mapping.get("email")
            raw_email = str(row[email_idx]).strip().lower() if (email_idx is not None and email_idx < len(row) and row[email_idx] is not None) else ""
            
            password_idx = col_mapping.get("password")
            custom_password = str(row[password_idx]).strip() if (password_idx is not None and password_idx < len(row) and row[password_idx] is not None) else ""
            
            if custom_password and custom_password != "None":
                hashed_user_password = get_password_hash(custom_password)
            else:
                hashed_user_password = hashed_default_password
            
            with db.begin_nested():
                from sqlalchemy import func, and_
                student = db.query(Student).filter(func.lower(Student.register_no) == func.lower(register_no)).first()
                
                clean_reg = register_no.replace(" ", "").strip()
                email = f"{clean_reg}@kce.ac.in"
                
                user = None
                if student and student.user_id:
                    user = db.query(User).filter(User.id == student.user_id).first()
                
                if not user:
                    user = db.query(User).filter(
                        (func.lower(User.username) == func.lower(register_no)) | 
                        (func.lower(User.email) == func.lower(email))
                    ).first()
                
                if not user:
                    user = User(
                        username=register_no,
                        email=email,
                        password_hash=hashed_user_password,
                        role="student",
                        is_active=True
                    )
                    db.add(user)
                    db.flush()
                else:
                    user.username = register_no
                    user.email = email
                    user.role = "student"
                    user.is_active = True
                    db.flush()
                
                if student:
                    student.name = name
                    student.user_id = user.id
                    student.department = department
                    student.year = year
                    student.section = section
                    if phone:
                        student.phone = phone
                    student.batch = batch
                    student.email = email
                    db.flush()
                    updated += 1
                else:
                    student = Student(
                        user_id=user.id,
                        register_no=register_no,
                        name=name,
                        email=email,
                        phone=phone,
                        department=department,
                        year=year,
                        section=section,
                        batch=batch
                    )
                    db.add(student)
                    db.flush()
                    inserted += 1
                
                up = db.query(UserProfile).filter(UserProfile.user_id == user.id).first()
                if not up:
                    up = UserProfile(
                        user_id=user.id,
                        full_name=name,
                        email=email,
                        phone=phone,
                        department=department,
                        location="Coimbatore"
                    )
                    db.add(up)
                else:
                    up.full_name = name
                    up.email = email
                    if phone:
                        up.phone = phone
                    up.department = department
                
                about = db.query(StudentAbout).filter(StudentAbout.student_id == student.id).first()
                if not about:
                    about = StudentAbout(
                        student_id=student.id,
                        headline="AI & DS Student | Java Full Stack Developer | Aspiring AI Engineer",
                        about_me=f"I am {name}, studying Artificial Intelligence & Data Science at Karpagam College of Engineering.",
                        career_objective="To build a strong career as an AI Engineer and Full Stack Developer.",
                        skills_json='["AI & Data Science", "Java", "React", "Full Stack Development", "Python", "DSA", "DBMS", "FastAPI", "PostgreSQL"]'
                    )
                    db.add(about)
                    
                cust = db.query(PortfolioCustomization).filter(PortfolioCustomization.student_id == student.id).first()
                if not cust:
                    cust = PortfolioCustomization(
                        student_id=student.id,
                        headline="AI & DS Student | Java Full Stack Developer | Aspiring AI Engineer",
                        about_me=f"I am {name}, studying Artificial Intelligence & Data Science at Karpagam College of Engineering.",
                        career_objective="To build a strong career as an AI Engineer and Full Stack Developer.",
                        skills_json='["AI & Data Science", "Java", "React", "Full Stack Development", "Python", "DSA", "DBMS", "FastAPI", "PostgreSQL"]',
                        theme="Dark Minimal",
                        section_visibility_json='{"showProjects":true,"showCertifications":true,"showAchievements":true,"showAcademicHighlights":true,"showContactLinks":true}',
                        resume_visibility=True
                    )
                    db.add(cust)
                    
                db.flush()
                
                # Auto Mentor Assignment logic
                m_email_idx = col_mapping.get("mentor_email")
                mentor_email = str(row[m_email_idx]).strip() if (m_email_idx is not None and m_email_idx < len(row) and row[m_email_idx] is not None) else ""
                
                mentor_to_assign = None
                if mentor_email:
                    mentor_to_assign = db.query(User).filter(User.email.ilike(mentor_email), User.role == "mentor").first()
                
                if not mentor_to_assign:
                    filter_conds = [
                        Student.department == department,
                        Student.section == section,
                        Student.batch == batch
                    ]
                    if year:
                        filter_conds.append(Student.year == year)
                    
                    matching_assignments = db.query(MentorAssignment.mentor_id).join(Student).filter(
                        and_(*filter_conds)
                    ).distinct().all()
                    
                    mentor_ids = [r[0] for r in matching_assignments]
                    if len(mentor_ids) == 1:
                        mentor_to_assign = db.query(User).filter(User.id == mentor_ids[0], User.role == "mentor").first()
                    elif len(mentor_ids) > 1:
                        print(f"[WARNING] Multiple mentors {mentor_ids} found for class Dept={department}, Yr={year}, Sec={section}, Batch={batch}. Skipping auto-assignment for student {register_no}.")
                
                if mentor_to_assign:
                    existing_assign = db.query(MentorAssignment).filter(
                        MentorAssignment.student_id == student.id,
                        MentorAssignment.mentor_id == mentor_to_assign.id
                    ).first()
                    if not existing_assign:
                        any_assign = db.query(MentorAssignment).filter(MentorAssignment.student_id == student.id).first()
                        if not any_assign:
                            new_assign = MentorAssignment(mentor_id=mentor_to_assign.id, student_id=student.id)
                            db.add(new_assign)
                            
                recalculate_student_analytics(db, student.id)
            
        except Exception as e:
            errors_list.append({
                "row": idx,
                "identifier": register_no if 'register_no' in locals() and register_no else "Row " + str(idx),
                "message": str(e)
            })
            skipped += 1
            
    try:
        db.commit()
        from app.services.cache_service import invalidate_all_caches
        invalidate_all_caches()
    except Exception as commit_err:
        db.rollback()
        errors_list.append({
            "row": "Finalization",
            "identifier": "Database Save",
            "message": f"Failed to commit batch: {str(commit_err)}"
        })
            
    return {
        "success": True,
        "type": "students",
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors_list
    }


# Endpoint: Faculty Upload
@router.post("/upload/faculty")
async def upload_faculty(
    file: UploadFile = File(...),
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    rows = parse_upload_file(file)
    if not rows or len(rows) < 2:
        return {
            "success": False,
            "type": "faculty",
            "inserted": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [{"row": 1, "identifier": "File", "message": "Spreadsheet is empty or missing data rows"}]
        }
        
    expected_cols = {
        "name": ["name", "faculty name", "faculty_name", "full name", "full_name"],
        "department": ["department", "dept", "branch"],
        "email": ["email", "email address", "email_address"],
        "phone": ["phone", "phone no", "phone_no", "phone number", "mobile", "mobile no", "mobile_no"],
        "designation": ["designation", "role", "title", "position"],
        "employee_id": ["employee_id", "employee id", "emp id", "emp_id", "faculty id", "faculty_id"],
        "password": ["password", "default password", "default_password", "pass"]
    }
    
    col_mapping = map_headers(rows[0], expected_cols)
    if "name" not in col_mapping:
        raise HTTPException(status_code=400, detail="Missing required column: name")
        
    inserted = 0
    updated = 0
    skipped = 0
    errors_list = []
    
    hashed_default_password = get_password_hash("Password123!")
    
    for idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            skipped += 1
            continue
            
        try:
            name_idx = col_mapping["name"]
            name = str(row[name_idx]).strip() if (name_idx < len(row) and row[name_idx] is not None) else ""
            if not name or name == "None":
                errors_list.append({"row": idx, "identifier": "Row " + str(idx), "message": "Faculty name is missing"})
                skipped += 1
                continue
                
            email_idx = col_mapping.get("email")
            email = str(row[email_idx]).strip().lower() if (email_idx is not None and email_idx < len(row) and row[email_idx] is not None) else ""
            
            emp_id_idx = col_mapping.get("employee_id")
            emp_id = str(row[emp_id_idx]).strip() if (emp_id_idx is not None and emp_id_idx < len(row) and row[emp_id_idx] is not None) else ""
            
            if not email:
                email = clean_and_generate_email(name, emp_id, db)
                
            dept_idx = col_mapping.get("department")
            department = str(row[dept_idx]).strip() if (dept_idx is not None and dept_idx < len(row) and row[dept_idx] is not None) else "AI & DS"
            
            phone_idx = col_mapping.get("phone")
            phone = str(row[phone_idx]).strip() if (phone_idx is not None and phone_idx < len(row) and row[phone_idx] is not None) else ""
            
            desig_idx = col_mapping.get("designation")
            designation = str(row[desig_idx]).strip() if (desig_idx is not None and desig_idx < len(row) and row[desig_idx] is not None) else "Assistant Professor"

            password_idx = col_mapping.get("password")
            custom_password = str(row[password_idx]).strip() if (password_idx is not None and password_idx < len(row) and row[password_idx] is not None) else ""
            
            if custom_password and custom_password != "None":
                hashed_user_password = get_password_hash(custom_password)
            else:
                hashed_user_password = hashed_default_password
            
            with db.begin_nested():
                user = db.query(User).filter(User.email == email).first()
                if user:
                    user.is_active = True
                    up = db.query(UserProfile).filter(UserProfile.user_id == user.id).first()
                    if up:
                        up.full_name = name
                        up.phone = phone
                        up.department = department
                    else:
                        up = UserProfile(user_id=user.id, full_name=name, email=email, phone=phone, department=department)
                        db.add(up)
                        
                    fp = db.query(FacultyProfile).filter(FacultyProfile.user_id == user.id).first()
                    if fp:
                        fp.name = name
                        fp.phone = phone
                        fp.department = department
                        fp.designation = designation
                    else:
                        fp = FacultyProfile(user_id=user.id, name=name, email=email, phone=phone, department=department, designation=designation)
                        db.add(fp)
                    db.flush()
                    updated += 1
                else:
                    user = User(
                        username=email,
                        email=email,
                        password_hash=hashed_user_password,
                        role="faculty"
                    )
                    db.add(user)
                    db.flush()
                    
                    up = UserProfile(
                        user_id=user.id,
                        full_name=name,
                        email=email,
                        phone=phone,
                        department=department,
                        location="Coimbatore"
                    )
                    db.add(up)
                    
                    fp = FacultyProfile(
                        user_id=user.id,
                        name=name,
                        email=email,
                        phone=phone,
                        department=department,
                        designation=designation
                    )
                    db.add(fp)
                    db.flush()
                    inserted += 1
                    
        except Exception as e:
            errors_list.append({
                "row": idx,
                "identifier": email if 'email' in locals() and email else "Row " + str(idx),
                "message": str(e)
            })
            skipped += 1
            
    try:
        db.commit()
    except Exception as commit_err:
        db.rollback()
        errors_list.append({
            "row": "Finalization",
            "identifier": "Database Save",
            "message": f"Failed to commit batch: {str(commit_err)}"
        })
            
    return {
        "success": True,
        "type": "faculty",
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors_list
    }


# Endpoint: Mentors Upload
@router.post("/upload/mentors")
async def upload_mentors(
    file: UploadFile = File(...),
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    rows = parse_upload_file(file)
    if not rows or len(rows) < 2:
        return {
            "success": False,
            "type": "mentors",
            "inserted": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [{"row": 1, "identifier": "File", "message": "Spreadsheet is empty or missing data rows"}]
        }
        
    expected_cols = {
        "name": ["name", "mentor name", "mentor_name", "full name", "full_name"],
        "department": ["department", "dept", "branch"],
        "email": ["email", "email address", "email_address"],
        "phone": ["phone", "phone no", "phone_no", "phone number", "mobile", "mobile no", "mobile_no"],
        "designation": ["designation", "role", "title", "position"],
        "employee_id": ["employee_id", "employee id", "emp id", "emp_id", "mentor id", "mentor_id"],
        "assigned_section": ["assigned_section", "assigned section", "section", "sec"],
        "assigned_batch": ["assigned_batch", "assigned batch", "batch", "class batch"],
        "password": ["password", "default password", "default_password", "pass"]
    }
    
    col_mapping = map_headers(rows[0], expected_cols)
    if "name" not in col_mapping:
        raise HTTPException(status_code=400, detail="Missing required column: name")
        
    inserted = 0
    updated = 0
    skipped = 0
    errors_list = []
    
    hashed_default_password = get_password_hash("Password123!")
    
    for idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            skipped += 1
            continue
            
        try:
            name_idx = col_mapping["name"]
            name = str(row[name_idx]).strip() if (name_idx < len(row) and row[name_idx] is not None) else ""
            if not name or name == "None":
                errors_list.append({"row": idx, "identifier": "Row " + str(idx), "message": "Mentor name is missing"})
                skipped += 1
                continue
                
            email_idx = col_mapping.get("email")
            email = str(row[email_idx]).strip().lower() if (email_idx is not None and email_idx < len(row) and row[email_idx] is not None) else ""
            
            emp_id_idx = col_mapping.get("employee_id")
            emp_id = str(row[emp_id_idx]).strip() if (emp_id_idx is not None and emp_id_idx < len(row) and row[emp_id_idx] is not None) else ""
            
            if not email:
                email = clean_and_generate_email(name, emp_id, db)
                
            dept_idx = col_mapping.get("department")
            department = str(row[dept_idx]).strip() if (dept_idx is not None and dept_idx < len(row) and row[dept_idx] is not None) else "AI & DS"
            
            phone_idx = col_mapping.get("phone")
            phone = str(row[phone_idx]).strip() if (phone_idx is not None and phone_idx < len(row) and row[phone_idx] is not None) else ""
            
            desig_idx = col_mapping.get("designation")
            designation = str(row[desig_idx]).strip() if (desig_idx is not None and desig_idx < len(row) and row[desig_idx] is not None) else "Mentor"

            password_idx = col_mapping.get("password")
            custom_password = str(row[password_idx]).strip() if (password_idx is not None and password_idx < len(row) and row[password_idx] is not None) else ""
            
            if custom_password and custom_password != "None":
                hashed_user_password = get_password_hash(custom_password)
            else:
                hashed_user_password = hashed_default_password
            
            with db.begin_nested():
                user = db.query(User).filter(User.email == email).first()
                if user:
                    user.is_active = True
                    up = db.query(UserProfile).filter(UserProfile.user_id == user.id).first()
                    if up:
                        up.full_name = name
                        up.phone = phone
                        up.department = department
                    else:
                        up = UserProfile(user_id=user.id, full_name=name, email=email, phone=phone, department=department)
                        db.add(up)
                        
                    fp = db.query(FacultyProfile).filter(FacultyProfile.user_id == user.id).first()
                    if fp:
                        fp.name = name
                        fp.phone = phone
                        fp.department = department
                        fp.designation = designation
                    else:
                        fp = FacultyProfile(user_id=user.id, name=name, email=email, phone=phone, department=department, designation=designation)
                        db.add(fp)
                        
                    db.flush()
                    updated += 1
                else:
                    user = User(
                        username=email,
                        email=email,
                        password_hash=hashed_user_password,
                        role="mentor"
                    )
                    db.add(user)
                    db.flush()
                    
                    up = UserProfile(
                        user_id=user.id,
                        full_name=name,
                        email=email,
                        phone=phone,
                        department=department,
                        location="Coimbatore"
                    )
                    db.add(up)
                    
                    fp = FacultyProfile(
                        user_id=user.id,
                        name=name,
                        email=email,
                        phone=phone,
                        department=department,
                        designation=designation
                    )
                    db.add(fp)
                    db.flush()
                    inserted += 1
                
                sec_idx = col_mapping.get("assigned_section")
                batch_idx = col_mapping.get("assigned_batch")
                
                assigned_sec = str(row[sec_idx]).strip() if (sec_idx is not None and sec_idx < len(row) and row[sec_idx] is not None) else ""
                assigned_bat = str(row[batch_idx]).strip() if (batch_idx is not None and batch_idx < len(row) and row[batch_idx] is not None) else ""
                
                if assigned_sec and assigned_bat:
                    students = db.query(Student).filter(
                        Student.department.ilike(department),
                        Student.section.ilike(assigned_sec),
                        Student.batch.ilike(assigned_bat)
                    ).all()
                    
                    for s in students:
                        existing_assign = db.query(MentorAssignment).filter(
                            MentorAssignment.mentor_id == user.id,
                            MentorAssignment.student_id == s.id
                        ).first()
                        if not existing_assign:
                            new_assign = MentorAssignment(mentor_id=user.id, student_id=s.id)
                            db.add(new_assign)
                            
        except Exception as e:
            errors_list.append({
                "row": idx,
                "identifier": email if 'email' in locals() and email else "Row " + str(idx),
                "message": str(e)
            })
            skipped += 1
            
    try:
        db.commit()
    except Exception as commit_err:
        db.rollback()
        errors_list.append({
            "row": "Finalization",
            "identifier": "Database Save",
            "message": f"Failed to commit batch: {str(commit_err)}"
        })
        
    return {
        "success": True,
        "type": "mentors",
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors_list
    }


@router.get("/debug/student/{register_no}")
async def admin_debug_student(
    register_no: str,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """Debug endpoint to verify a student profile and its user linking."""
    from sqlalchemy import func
    from app.models.student import MentorAssignment
    
    student = db.query(Student).filter(func.lower(Student.register_no) == func.lower(register_no)).first()
    
    if not student:
        return {
            "student_exists": False,
            "student_id": None,
            "register_no": register_no,
            "student_user_id": None,
            "user_exists": False,
            "user_id": None,
            "user_username": None,
            "user_email": None,
            "user_role": None,
            "user_is_active": None,
            "mentor_assignments_count": 0
        }
        
    user = db.query(User).filter(User.id == student.user_id).first()
    assignments_count = db.query(MentorAssignment).filter(MentorAssignment.student_id == student.id).count()
    
    return {
        "student_exists": True,
        "student_id": student.id,
        "register_no": student.register_no,
        "student_user_id": student.user_id,
        "user_exists": user is not None,
        "user_id": user.id if user else None,
        "user_username": user.username if user else None,
        "user_email": user.email if user else None,
        "user_role": user.role if user else None,
        "user_is_active": user.is_active if user else None,
        "mentor_assignments_count": assignments_count
    }


@router.get("/debug/scores/{register_no}")
async def admin_debug_scores(
    register_no: str,
    current_user: User = Depends(RoleRequired(["admin", "faculty", "mentor"])),
    db: Session = Depends(get_db)
):
    """Debug endpoint to check score counts, domain averages, and analytics state for a student."""
    from sqlalchemy import func
    from app.models.score import AssessmentScore, StudentAnalytics
    from app.utils.domain_utils import normalize_domain

    clean_reg = register_no.strip()
    if clean_reg.endswith(".0"):
        clean_reg = clean_reg[:-2]

    student = db.query(Student).filter(func.lower(Student.register_no) == clean_reg.lower()).first()

    if not student:
        return {
            "register_no": register_no,
            "student_found": False,
            "student_id": None,
            "scores_count": 0,
            "categories": {
                "DSA": 0.0,
                "DBMS": 0.0,
                "FullStack": 0.0,
                "Aptitude": 0.0,
                "Coding": 0.0,
                "Academic": 0.0
            },
            "analytics_exists": False,
            "overall_score": None
        }

    scores = db.query(AssessmentScore).filter(AssessmentScore.student_id == student.id).all()
    analytics = db.query(StudentAnalytics).filter(StudentAnalytics.student_id == student.id).first()

    categories = ["DSA", "DBMS", "FullStack", "Aptitude", "Coding", "Academic"]
    cat_map = {cat: [] for cat in categories}
    for sc in scores:
        norm_c = normalize_domain(sc.category)
        if norm_c and norm_c in cat_map:
            cat_map[norm_c].append(sc.percentage)

    cat_averages = {}
    for cat in categories:
        vals = cat_map[cat]
        cat_averages[cat] = round(sum(vals) / len(vals), 2) if vals else 0.0

    overall = analytics.overall_score if (analytics and analytics.total_assessments > 0) else None

    return {
        "register_no": student.register_no,
        "student_found": True,
        "student_id": student.id,
        "scores_count": len(scores),
        "categories": cat_averages,
        "analytics_exists": analytics is not None,
        "overall_score": overall
    }


from pydantic import BaseModel, Field

class AssignStudentsPayload(BaseModel):
    mentor_email: Optional[str] = None
    mentor_id: Optional[int] = None
    register_numbers: List[str] = Field(default_factory=list)

class AssignAllStudentsPayload(BaseModel):
    mentor_email: Optional[str] = None
    mentor_id: Optional[int] = None


@router.post("/mentors/assign-all-students")
@router.post("/mentors/{mentor_id}/assign-all-students")
async def admin_assign_all_students_to_mentor(
    payload: Optional[AssignAllStudentsPayload] = None,
    mentor_id: Optional[int] = None,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """Assigns ALL available real uploaded students (excluding 22AD demo students) to a mentor."""
    from sqlalchemy import func
    from app.models.student import MentorAssignment

    target_mentor_id = mentor_id or (payload.mentor_id if payload else None)
    target_email = payload.mentor_email if payload else None

    mentor_user = None
    if target_email:
        mentor_user = db.query(User).filter(func.lower(User.email) == func.lower(target_email.strip())).first()

    if not mentor_user and target_mentor_id:
        mentor_user = db.query(User).filter(User.id == target_mentor_id).first()

    if not mentor_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Mentor not found with provided email '{target_email}' or ID '{target_mentor_id}'"
        )

    all_students = db.query(Student).all()
    demo_students = [s for s in all_students if s.register_no and s.register_no.strip().lower().startswith("22ad")]
    real_students = [s for s in all_students if not (s.register_no and s.register_no.strip().lower().startswith("22ad"))]
    excluded_demo_students = len(demo_students)

    assigned_count = 0
    already_assigned_count = 0

    existing_assignments = db.query(MentorAssignment).filter(MentorAssignment.mentor_id == mentor_user.id).all()
    existing_student_ids = set(a.student_id for a in existing_assignments)

    for student in real_students:
        if student.id in existing_student_ids:
            already_assigned_count += 1
        else:
            assignment = MentorAssignment(
                mentor_id=mentor_user.id,
                student_id=student.id
            )
            db.add(assignment)
            assigned_count += 1
            existing_student_ids.add(student.id)

    db.commit()

    from app.services.cache_service import invalidate_all_caches
    invalidate_all_caches()

    return {
        "success": True,
        "mentor_email": mentor_user.email,
        "assigned": assigned_count,
        "already_assigned": already_assigned_count,
        "total_real_students": len(real_students),
        "excluded_demo_students": excluded_demo_students,
        "errors": []
    }


@router.post("/mentors/assign-students")
@router.post("/mentors/{mentor_id}/assign-students")
async def admin_assign_students_to_mentor(
    payload: AssignStudentsPayload,
    mentor_id: Optional[int] = None,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """Assigns students to a mentor by register numbers."""
    from sqlalchemy import func
    from app.models.student import MentorAssignment

    target_mentor_id = mentor_id or payload.mentor_id
    target_email = payload.mentor_email

    mentor_user = None
    if target_email:
        mentor_user = db.query(User).filter(func.lower(User.email) == func.lower(target_email.strip())).first()

    if not mentor_user and target_mentor_id:
        mentor_user = db.query(User).filter(User.id == target_mentor_id).first()

    if not mentor_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Mentor not found with provided email '{target_email}' or ID '{target_mentor_id}'"
        )

    # Process register numbers
    raw_regs = payload.register_numbers
    parsed_regs = []
    for item in raw_regs:
        if not item:
            continue
        for line in str(item).replace(",", "\n").splitlines():
            cleaned = line.strip()
            if cleaned and cleaned not in parsed_regs:
                parsed_regs.append(cleaned)

    assigned_count = 0
    already_assigned_count = 0
    not_found = []
    errors = []

    for reg in parsed_regs:
        student = db.query(Student).filter(func.lower(Student.register_no) == func.lower(reg)).first()
        if not student:
            not_found.append(reg)
            continue

        existing = db.query(MentorAssignment).filter(
            MentorAssignment.mentor_id == mentor_user.id,
            MentorAssignment.student_id == student.id
        ).first()

        if existing:
            already_assigned_count += 1
        else:
            assignment = MentorAssignment(
                mentor_id=mentor_user.id,
                student_id=student.id
            )
            db.add(assignment)
            assigned_count += 1

    db.commit()

    from app.services.cache_service import invalidate_all_caches
    invalidate_all_caches()

    return {
        "success": True,
        "mentor_email": mentor_user.email,
        "assigned": assigned_count,
        "already_assigned": already_assigned_count,
        "not_found": not_found,
        "errors": errors
    }


@router.post("/mentors/upload-assignments")
async def admin_upload_mentor_assignments(
    file: UploadFile = File(...),
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """Uploads Excel sheet to bulk assign students to mentors."""
    from sqlalchemy import func
    from app.models.student import MentorAssignment

    try:
        file_bytes = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file format: {str(e)}")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows or len(rows) < 2:
        return {
            "success": False,
            "assigned": 0,
            "already_assigned": 0,
            "not_found": [],
            "errors": ["Excel file is empty or has no data rows"]
        }

    headers = [str(c).strip().lower().replace("_", " ") if c is not None else "" for c in rows[0]]

    email_col = None
    reg_col = None

    for idx, h in enumerate(headers):
        if h in ["mentor email", "mentor_email", "email", "mentor"]:
            email_col = idx
        elif h in ["register no", "register_no", "reg no", "reg_no", "register number", "reg number"]:
            reg_col = idx

    if email_col is None or reg_col is None:
        raise HTTPException(
            status_code=400,
            detail="Missing required columns in Excel: 'Mentor Email' and 'Register No'"
        )

    all_users = db.query(User).all()
    user_map = {u.email.strip().lower(): u for u in all_users if u.email}

    all_students = db.query(Student).all()
    student_map = {s.register_no.strip().lower(): s for s in all_students if s.register_no}

    assigned_count = 0
    already_assigned_count = 0
    not_found = []
    errors = []

    for r_idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        raw_email = str(row[email_col]).strip() if row[email_col] is not None else ""
        raw_reg = str(row[reg_col]).strip() if row[reg_col] is not None else ""
        if raw_reg.endswith(".0"):
            raw_reg = raw_reg[:-2]

        if not raw_email or not raw_reg:
            continue

        mentor_user = user_map.get(raw_email.lower())
        if not mentor_user:
            errors.append(f"Row {r_idx}: Mentor with email '{raw_email}' not found")
            continue

        student = student_map.get(raw_reg.lower())
        if not student:
            not_found.append(raw_reg)
            continue

        existing = db.query(MentorAssignment).filter(
            MentorAssignment.mentor_id == mentor_user.id,
            MentorAssignment.student_id == student.id
        ).first()

        if existing:
            already_assigned_count += 1
        else:
            assignment = MentorAssignment(
                mentor_id=mentor_user.id,
                student_id=student.id
            )
            db.add(assignment)
            assigned_count += 1

    db.commit()

    from app.services.cache_service import invalidate_all_caches
    invalidate_all_caches()

    return {
        "success": True,
        "assigned": assigned_count,
        "already_assigned": already_assigned_count,
        "not_found": not_found,
        "errors": errors
    }


@router.get("/debug/mentor-assignments/{mentor_email}")
async def admin_debug_mentor_assignments(
    mentor_email: str,
    current_user: User = Depends(RoleRequired(["admin", "faculty", "mentor"])),
    db: Session = Depends(get_db)
):
    """Debug endpoint to check mentor student assignment records."""
    from sqlalchemy import func
    from app.models.student import MentorAssignment

    mentor_user = db.query(User).filter(func.lower(User.email) == func.lower(mentor_email.strip())).first()
    if not mentor_user:
        return {
            "mentor_found": False,
            "mentor_email": mentor_email,
            "assigned_count": 0,
            "assigned_register_numbers": []
        }

    assignments = db.query(MentorAssignment, Student)\
        .join(Student, MentorAssignment.student_id == Student.id)\
        .filter(MentorAssignment.mentor_id == mentor_user.id).all()

    assigned_regs = [s.register_no for a, s in assignments]

    return {
        "mentor_found": True,
        "mentor_email": mentor_user.email,
        "assigned_count": len(assigned_regs),
        "assigned_register_numbers": assigned_regs
    }


@router.get("/debug/mentor-student-access/{mentor_email}/{register_no}")
async def admin_debug_mentor_student_access(
    mentor_email: str,
    register_no: str,
    db: Session = Depends(get_db)
):
    """Debug endpoint to inspect mentor student access status and assignment row resolution."""
    from sqlalchemy import func
    from app.models.student import MentorAssignment

    mentor_user = db.query(User).filter(func.lower(User.email) == func.lower(mentor_email.strip())).first()
    if not mentor_user:
        return {
            "mentor_found": False,
            "mentor_id": None,
            "mentor_email": mentor_email,
            "student_found": False,
            "student_id": None,
            "register_no": register_no,
            "is_assigned": False,
            "assignment_row_found": False,
            "reason": "mentor not found"
        }

    clean_reg = str(register_no).strip()
    student = db.query(Student).filter(func.lower(Student.register_no) == clean_reg.lower()).first()
    if not student and clean_reg.isdigit():
        student = db.query(Student).filter(Student.id == int(clean_reg)).first()

    if not student:
        return {
            "mentor_found": True,
            "mentor_id": mentor_user.id,
            "mentor_email": mentor_user.email,
            "student_found": False,
            "student_id": None,
            "register_no": register_no,
            "is_assigned": False,
            "assignment_row_found": False,
            "reason": "student not found"
        }

    assignment = db.query(MentorAssignment).filter(
        MentorAssignment.mentor_id == mentor_user.id,
        MentorAssignment.student_id == student.id
    ).first()

    is_assigned = assignment is not None

    return {
        "mentor_found": True,
        "mentor_id": mentor_user.id,
        "mentor_email": mentor_user.email,
        "student_found": True,
        "student_id": student.id,
        "register_no": student.register_no,
        "is_assigned": is_assigned,
        "assignment_row_found": is_assigned,
        "reason": "assigned through mentor_assignments" if is_assigned else "assignment missing"
    }


@router.post("/upload/scores", response_model=UploadScoresResponse)
@router.post("/scores/upload", response_model=UploadScoresResponse)
async def admin_upload_scores(
    file: UploadFile = File(...),
    current_user: User = Depends(RoleRequired(["admin", "faculty", "mentor"])),
    db: Session = Depends(get_db)
):
    """Admin score upload handler alias."""
    file_bytes = await file.read()
    return process_scores_excel(db, file_bytes, current_user.id)


@router.post("/students/{register_no}/reset-password")
async def admin_reset_student_password(
    register_no: str,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """
    Admin resets a student's password to a temporary password.
    Temporary password is shown exactly once in the response.
    It is hashed with bcrypt before saving.
    """
    import secrets
    import string
    from app.models.otp_models import PasswordResetLog
    from app.models.student import Student
    from app.utils.security import get_password_hash
    from sqlalchemy import func

    # Find student
    student = db.query(Student).filter(func.lower(Student.register_no) == register_no.lower()).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Find user
    user = db.query(User).filter(User.id == student.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User account not found for student")

    # Generate temporary password
    alphabet = string.ascii_letters + string.digits
    temp_pass = ''.join(secrets.choice(alphabet) for i in range(10))

    # Hash and save password
    user.password_hash = get_password_hash(temp_pass)
    
    # Log reset
    log = PasswordResetLog(
        user_id=user.id,
        admin_id=current_user.id,
        email=user.email,
        register_no=student.register_no,
        role=user.role,
        action="admin_password_reset",
        status="success",
        message=f"Password reset by administrator {current_user.email}"
    )
    db.add(log)
    db.commit()

    return {
        "success": True,
        "temporary_password": temp_pass,
        "message": "Password reset successfully. Copy this password now. It will not be shown again."
    }


@router.post("/impersonate/{register_no}")
async def admin_impersonate_student(
    register_no: str,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """
    Admin impersonates a student and logs in as that student.
    Generates a 15-minute token with student role, containing impersonation details.
    """
    from datetime import timedelta
    from app.models.otp_models import PasswordResetLog
    from app.models.student import Student
    from app.services.auth_service import create_user_auth_payload
    from app.utils.security import create_access_token, create_refresh_token
    from sqlalchemy import func

    # Find student
    student = db.query(Student).filter(func.lower(Student.register_no) == register_no.lower()).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Find user
    user = db.query(User).filter(User.id == student.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User account not found for student")

    # Generate student token payload with impersonation claims
    # Expiry is 15 minutes as per safety requirements
    token_data = {
        "sub": str(user.id),
        "email": user.email,
        "role": user.role,
        "impersonation": True,
        "impersonated_by_admin_id": current_user.id,
        "original_admin_email": current_user.email
    }

    # Generate token with 15 min expiry
    access_token = create_access_token(data=token_data, expires_delta=timedelta(minutes=15))
    refresh_token = create_refresh_token(data=token_data, expires_delta=timedelta(minutes=15))
    user_payload = create_user_auth_payload(user, db)

    # Log impersonation
    log = PasswordResetLog(
        user_id=user.id,
        admin_id=current_user.id,
        email=user.email,
        register_no=student.register_no,
        role=user.role,
        action="admin_impersonation_started",
        status="success",
        message=f"Administrator {current_user.email} started impersonation session"
    )
    db.add(log)
    db.commit()

    return {
        "success": True,
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "expires_in_minutes": 15,
        "impersonation": True,
        "student": {
            "register_no": student.register_no,
            "name": student.name
        },
        "user": user_payload
    }


@router.get("/password-reset-logs")
async def get_password_reset_logs(
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """
    Get audit logs for password resets and impersonation.
    """
    from app.models.otp_models import PasswordResetLog
    
    logs = db.query(PasswordResetLog).order_by(PasswordResetLog.created_at.desc()).limit(100).all()
    
    result = []
    for log in logs:
        result.append({
            "id": log.id,
            "user_id": log.user_id,
            "admin_id": log.admin_id,
            "email": log.email,
            "register_no": log.register_no,
            "role": log.role,
            "action": log.action,
            "status": log.status,
            "message": log.message,
            "created_at": log.created_at.isoformat() if log.created_at else None
        })
    return result


@router.post("/students/update-register-emails")
async def update_student_register_emails(
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """
    Updates all existing student user emails to register_no@kce.ac.in.
    Checks for duplicates/conflicts first and skips them, reporting conflicts.
    """
    from app.models.student import Student
    from app.models.user import User, UserProfile
    from sqlalchemy import func
    
    # 1. Fetch all students
    students = db.query(Student).all()
    
    # 2. Check for duplicate register numbers in the students table
    reg_counts = {}
    for s in students:
        if s.register_no:
            reg = s.register_no.strip().replace(" ", "")
            reg_counts[reg] = reg_counts.get(reg, 0) + 1
            
    conflicts = []
    updated_count = 0
    skipped_count = 0
    errors = []
    
    for student in students:
        if not student.register_no:
            skipped_count += 1
            continue
            
        clean_reg = student.register_no.strip().replace(" ", "")
        
        # Check if this register number is duplicate in the students table
        if reg_counts.get(clean_reg, 0) > 1:
            conflicts.append({
                "register_no": student.register_no,
                "reason": "Duplicate register number found in database"
            })
            skipped_count += 1
            continue
            
        if not student.user_id:
            skipped_count += 1
            continue
            
        user = db.query(User).filter(User.id == student.user_id).first()
        if not user:
            skipped_count += 1
            continue
            
        if user.role != "student":
            skipped_count += 1
            continue
            
        new_email = f"{clean_reg}@kce.ac.in"
        
        # Check if the email is already set correctly
        if user.email == new_email and user.username == clean_reg:
            # Update student.email and UserProfile just in case they are out of sync
            student.email = new_email
            up = db.query(UserProfile).filter(UserProfile.user_id == user.id).first()
            if up:
                up.email = new_email
            skipped_count += 1
            continue
            
        # Check if another user already has this email (email conflict)
        existing_with_email = db.query(User).filter(
            (func.lower(User.email) == new_email.lower()) & (User.id != user.id)
        ).first()
        if existing_with_email:
            conflicts.append({
                "register_no": student.register_no,
                "reason": f"Email conflict: email {new_email} is already used by another user (ID: {existing_with_email.id})"
            })
            skipped_count += 1
            continue
            
        try:
            # Update user
            user.email = new_email
            user.username = clean_reg
            
            # Update student
            student.email = new_email
            
            # Update UserProfile
            up = db.query(UserProfile).filter(UserProfile.user_id == user.id).first()
            if up:
                up.email = new_email
                
            db.flush()
            updated_count += 1
        except Exception as e:
            db.rollback()
            errors.append({
                "register_no": student.register_no,
                "error": str(e)
            })
            skipped_count += 1
            
    db.commit()
    
    return {
        "success": True,
        "updated": updated_count,
        "skipped": skipped_count,
        "conflicts": conflicts,
        "errors": errors
    }


from pydantic import BaseModel

class TestEmailRequest(BaseModel):
    to_email: str

@router.get("/debug/email-config")
async def debug_email_config(
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """
    Returns diagnostic config status for SMTP without exposing sensitive credentials.
    """
    import os
    from app.config import settings
    
    smtp_host = os.environ.get("SMTP_HOST") or os.environ.get("MAIL_SERVER") or settings.SMTP_HOST
    port_val = os.environ.get("SMTP_PORT") or os.environ.get("MAIL_PORT") or settings.SMTP_PORT
    smtp_port = int(port_val) if port_val else None
    smtp_user = os.environ.get("SMTP_USERNAME") or os.environ.get("MAIL_USERNAME") or os.environ.get("SMTP_USER") or settings.SMTP_USER
    smtp_password = os.environ.get("SMTP_PASSWORD") or os.environ.get("MAIL_PASSWORD") or settings.SMTP_PASSWORD
    smtp_from = os.environ.get("SMTP_FROM_EMAIL") or os.environ.get("MAIL_FROM") or os.environ.get("SMTP_FROM") or settings.SMTP_FROM
    
    smtp_configured = all([smtp_host, smtp_port, smtp_user, smtp_password, smtp_from])
    
    return {
        "smtp_configured": bool(smtp_configured),
        "smtp_host_present": bool(smtp_host),
        "smtp_port_present": bool(smtp_port),
        "smtp_username_present": bool(smtp_user),
        "smtp_password_present": bool(smtp_password),
        "from_email_present": bool(smtp_from)
    }

@router.post("/debug/send-test-email")
async def debug_send_test_email(
    payload: TestEmailRequest,
    current_user: User = Depends(RoleRequired(["admin"])),
    db: Session = Depends(get_db)
):
    """
    Sends a test email to the specified address.
    """
    import os
    import smtplib
    from email.mime.text import MIMEText
    from app.config import settings
    
    smtp_host = os.environ.get("SMTP_HOST") or os.environ.get("MAIL_SERVER") or settings.SMTP_HOST
    port_val = os.environ.get("SMTP_PORT") or os.environ.get("MAIL_PORT") or settings.SMTP_PORT
    smtp_port = int(port_val) if port_val else None
    smtp_user = os.environ.get("SMTP_USERNAME") or os.environ.get("MAIL_USERNAME") or os.environ.get("SMTP_USER") or settings.SMTP_USER
    smtp_password = os.environ.get("SMTP_PASSWORD") or os.environ.get("MAIL_PASSWORD") or settings.SMTP_PASSWORD
    smtp_from = os.environ.get("SMTP_FROM_EMAIL") or os.environ.get("MAIL_FROM") or os.environ.get("SMTP_FROM") or settings.SMTP_FROM
    
    smtp_tls_val = os.environ.get("SMTP_TLS") or os.environ.get("MAIL_TLS") or "true"
    use_tls = smtp_tls_val.lower() == "true"
    
    if not all([smtp_host, smtp_port, smtp_user, smtp_password, smtp_from]):
        return {
            "success": False,
            "message": "Email sending failed",
            "error_type": "ConfigurationError: SMTP variables are not fully configured."
        }
        
    try:
        msg = MIMEText("This is a diagnostic test email from the Student360 platform.")
        msg["Subject"] = "Student360 SMTP Diagnostic Test"
        msg["From"] = smtp_from
        msg["To"] = payload.to_email
        
        server = smtplib.SMTP(smtp_host, int(smtp_port))
        if use_tls:
            server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
        server.quit()
        
        return {
            "success": True,
            "message": "Test email sent"
        }
    except Exception as e:
        return {
            "success": False,
            "message": "Email sending failed",
            "error_type": type(e).__name__
        }

